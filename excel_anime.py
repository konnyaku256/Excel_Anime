import cv2
import numpy as np
from string import ascii_uppercase as AtoZ
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

image_path = './video/'

IMG_CNT = 1384 # 画像の枚数
IMG_W = 320   # 画像の幅
IMG_H = 180   # 画像の高さ

print('init')

'''
エクセルの初期設定
ビットマップ化したものを各セルに塗りつぶすためにセルの形を変形
'''

# エクセルファイル読み込み
wb = load_workbook(filename='kemofure.xlsx')
# アクティブなシートを取り出す
ws = wb.active

# エクセルの列名リスト作成
col_name = list(AtoZ)          # 'A'～'Z'
for i in range(IMG_W - 26):    # 'AA'～
  col_name.append(AtoZ[i // 26] + AtoZ[i % 26])

for col in range(IMG_W):
  # 列の幅を変更
  ws.column_dimensions[col_name[col]].width = 0.3

  for row in range(IMG_H * IMG_CNT):
    # 行の高さを変更
    ws.row_dimensions[row].height = 1.5


'''
各セルの塗りつぶし
'''

next_row = 0
count = 0

for i in range(IMG_CNT):
    print('fill', i)
    
    filename_i = str(image_path) + str(i) + '.jpg'
    print(filename_i)
    image = cv2.imread(filename_i)
    
    for row in range(IMG_H):
      for col in range(IMG_W):
        # 赤、緑、青の値を青、緑、赤の順に16進数2桁ずつの文字列へ変換
        red = image.item(row, col, 2)
        green = image.item(row, col, 1)
        blue = image.item(row, col, 0)
        color = '%02x%02x%02x' % (red, green, blue)

        # 塗りつぶすセル名を取得
        cell_name = col_name[col] + str(row + 1 + next_row)
        # セル名をセット
        cell = ws[cell_name]
        # 塗りつぶし
        cell.fill = PatternFill(patternType='solid', fgColor=color)
      
      count += 1

    next_row = count

wb.save('kemofure.xlsx')
print('done')